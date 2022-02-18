#Data Driven testing
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions

wb=openpyxl.load_workbook("F:/selenium_pgms_BhanuSir/actitime2.xlsx")
sheet=wb["login"]
rc=sheet.max_row
for i in range(2,rc+1):
    print(i)
    un=sheet.cell(i,1).value
    print("User name is",un)
    pw=sheet.cell(i,2).value
    print("Password is",pw)

    driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(4)
    driver.get("https://demo.actitime.com/login.do")
    time.sleep(2)
    driver.find_element(By.ID,"username").send_keys(un)
    driver.find_element(By.NAME,"pwd").send_keys(pw)
    time.sleep(2)
    driver.find_element(By.XPATH,"//div[.='Login ']").click()

    #get title of home page
    wait=WebDriverWait(driver,8)
    wait.until(expected_conditions.title_contains("Enter"))
    sheet.cell(i,3).value=driver.title
    driver.quit()


wb.save("F:/selenium_pgms_BhanuSir/actitime_res.xlsx")
wb.close()
