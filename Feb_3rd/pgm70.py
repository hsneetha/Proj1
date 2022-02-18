#READING DATA FROM EXCEL
import openpyxl



# #1. open the excel file(workbook)
# wb=openpyxl.load_workbook("F:\selenium_pgms_BhanuSir\Book1.xlsx")
# #2. goto required sheet
# sheet1=wb['Sheet1']
# #3. get the value present in required cell
# for i in range(1,4):
#     for j in range(1,4):
#         v=sheet1.cell(i,j).value
#         print(v)
#
# #5. close the xl file
# wb.close()

#to print in matrix format
#1. open the excel file(workbook)
wb=openpyxl.load_workbook("F:\selenium_pgms_BhanuSir\Book1.xlsx")
#2. goto required sheet
sheet1=wb['Sheet1']
#3. get the value present in required cell
for i in range(1,4):
    for j in range(1,4):
        v=sheet1.cell(i,j).value
        print(v,end=' ')
    print()

#5. close the xl file
wb.close()
