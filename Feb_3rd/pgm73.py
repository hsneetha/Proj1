import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions

wb=openpyxl.load_workbook("F:/selenium_pgms_BhanuSir/actitime.xlsx")
sheet=wb["login"]
un=sheet.cell(2,1).value
print("User name is",un)
pw=sheet.cell(2,2).value
print("Password is",pw)

driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.implicitly_wait(4)
driver.get("https://demo.actitime.com/login.do")

driver.find_element(By.ID,"username").send_keys(un)
driver.find_element(By.NAME,"pwd").send_keys(pw)
driver.find_element(By.XPATH,"//div[.='Login ']").click()

#get title of home page
wait=WebDriverWait(driver,8)
wait.until(expected_conditions.title_contains("Enter"))
sheet.cell(2,3).value=driver.title

wb.save("F:/selenium_pgms_BhanuSir/actitime.xlsx")

wb.close()
driver.quit()
