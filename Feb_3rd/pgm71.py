#to print in matrix format
#READING DATA FROM EXCEL
import openpyxl

#1. open the excel file(workbook)
wb=openpyxl.load_workbook("F:\selenium_pgms_BhanuSir\Book1.xlsx")
#2. goto required sheet
sheet1=wb['Sheet1']
# to count the no of rows in sheet
rc=sheet1.max_row
print("Row count",rc)
# to count the no of columns in sheet
cc=sheet1.max_column
print("COulmn Count",cc)
#3. get the value present in required cell
for i in range(1,rc+1):
    for j in range(1,cc+1):
        v=sheet1.cell(i,j).value
        print(v,end=' ')
    print()

#5. close the xl file